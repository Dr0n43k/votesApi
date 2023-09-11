from django.contrib import admin
from openpyxl.workbook import Workbook
from .models import Person, Voting, PersonVotes
from djangoProject import settings
import smtplib
import email.mime.application
import email.mime.multipart
import email.mime.text
import email

class PersonAdmin(admin.ModelAdmin):
    readonly_fields = ['img_preview']


class ReportsAdmin(admin.ModelAdmin):
    actions = ['create_report']

    @admin.action(description='create report')
    def create_report(self, request, queryset):
        email_from = ""     # enter your email
        email_from_password = ""    # enter your email password
        email_to = request.user.email
        for j in range(len(queryset)):
            person = PersonVotes.objects.filter(voting=queryset[j])
            report = Workbook()
            report_sheet = report.active
            report_sheet.title = queryset[j].title
            report_sheet.cell(1, 1, "Имя участника")
            report_sheet.cell(1, 2, "Количество голосов")
            for i in range(len(person)):
                report_sheet.cell(i+2, 1, person[i].person.fullname)
                report_sheet.cell(i+2, 2, person[i].votes)
            report.save(f"{settings.BASE_DIR}/reports/{queryset[j].title}.xlsx")
            msg = email.mime.multipart.MIMEMultipart()
            msg['Subject'] = f'Отчет по голосованию {queryset[j].title}'
            msg['From'] = email_from
            msg['To'] = email_to
            body = email.mime.text.MIMEText(f"Отчет по голосованию {queryset[j].title}")
            msg.attach(body)
            fp = open(f"{settings.BASE_DIR}/reports/{queryset[j].title}.xlsx", 'rb')
            att = email.mime.application.MIMEApplication(fp.read(), _subtype="xlsx")
            fp.close()
            att.add_header('Content-Disposition', 'attachment', filename=queryset[j].title+".xlsx")
            msg.attach(att)
            s = smtplib.SMTP('smtp.gmail.com')
            s.starttls()
            s.login(email_from, email_from_password)
            s.sendmail(email_from, [email_to], msg.as_string())
            s.quit()

admin.site.register(Person, PersonAdmin)
admin.site.register(Voting, ReportsAdmin)
admin.site.register(PersonVotes)


